import openpyxl

wb = openpyxl.load_workbook('豆瓣top250电影.xlsx')
ws = wb['Sheet Copy']
for row in ws.rows:
    if row[0].value == "阿甘正传":
        row[1].value = 9.8
    elif row[0].value == "这个杀手不太冷":
        ws[f'B{row[1].row}'] = 9.6
    elif row[0].value == "肖申克的救赎":
        ws[f'B{row[1].row}'] = 9.7

wb.save('豆瓣top250电影.xlsx')