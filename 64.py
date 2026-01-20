import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

ws['A1'].number_format = "[RED]+#,###.00;[GREEN]-#,###.00"
ws["A1"] = 99

ws['A2'].number_format = "[RED]+#,###.00;[GREEN]-#,###.00"
ws["A2"] = -99

ws['A3'].number_format = "[RED];[GREEN];[BLUE];[YELLOW]"
ws['A3'] = 0

ws['A4'].number_format = "[RED];[GREEN];[BLUE];[YELLOW]"
ws['A4'] = 'FishC'

ws['A5'].number_format = "[=1]男;[=0]女"
ws['A5'] = 0

ws['A6'].number_format = "[=1]男;[=0]女"
ws['A6'] = 1

ws['A7'].number_format = "[=1]男;[=0]女"
ws['A7'] = 2

ws['A8'].number_format = "[<60][RED]不及格;[>=60][GREEN]及格"
ws['A8'] = 58

ws['A9'].number_format = "[<60][RED]不及格;[>=60][GREEN]及格"
ws['A9'] = 68

wb.save("test.xlsx")