from openpyxl import load_workbook
from openpyxl.styles import Alignment

wb = load_workbook("test.xlsx")
ws = wb['Sheet']
center_alignment = Alignment(horizontal='center')

for row in ws.iter_rows(min_col=2,min_row=2,max_col=6,max_row=5):
    ws[row[4].coordinate] = f'=IF({row[3].coordinate}<250,"B","A")'

    ws[row[4].coordinate].alignment = center_alignment

wb.save("test.xlsx")